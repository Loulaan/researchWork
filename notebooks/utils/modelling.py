import numpy as np
import rpy2.robjects as robjects
import rpy2.robjects.numpy2ri
from rpy2.robjects.packages import importr
import openpyxl

from utils.hmatr import Hmatr

rpy2.robjects.numpy2ri.activate()
utils = importr('utils')
utils.chooseCRANmirror(ind=1)
utils.install_packages('Rssa')
rssa = importr('Rssa')


def measureStatistics(func, Q, tail):
        return np.max(func[:(Q-tail)]), np.quantile(func[:(Q-tail)], 0.95)


def modelling(dictSeries:dict, iterNum:int, N:int, B:int, T:int, Q:int, L:int, r:int, method:str, destFile:str):
    '''
    :param dict dictSeries: The dictionary where key is the type of series and value is a series. Example: { 'Permanent': [x_1, ..., x_N] }.
    :param int iterNum: Number of iterations for modelling.
    :param int N: The len of series.
    :param int B: The len of base subseries.
    :param int T: The len of test subseries.
    :param int Q: The point of perturbation.
    :param int L: The window len.
    :param int r: Number of eigen vectors.
    :param str method: SVD method.
    :param str destFile: Name of the file for saving results.
    '''

    try:
        wb = openpyxl.load_workbook(filename = destFile)
        sheet = wb['Modelling']
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Modelling"
        sheet = wb['Modelling']

    for num, typeV in enumerate(dictSeries.keys()):
        statsRow = []
        statsCol = []
        statsSym = []
        statsDiag = []
        
        for i in range(iterNum):
            eps = np.random.randn(N) * 0.5
            
            seriesNoise = dictSeries[typeV] + eps
            hm = Hmatr(seriesNoise, B, T, L, neig=r, svdMethod=method)
            statsRow.append(measureStatistics(hm.getRow(), Q, hm.T))
            statsCol.append(measureStatistics(hm.getCol(), Q, hm.B))
            statsSym.append(measureStatistics(hm.getSym(), Q, hm.T))
            statsDiag.append(measureStatistics(hm.getDiag(), Q, hm.B + hm.T + 1))

        sheet.cell(row=num*4 + 1, column=1).value = typeV    
        sheet.cell(row=num*4 + 2, column=1).value = 'meanMax'
        sheet.cell(row=num*4 + 3, column=1).value = '95 procentile'


        sheet.cell(row=num*4 + 1, column=2).value = 'row'
        i = 2
        for rec in np.mean(statsRow, axis=0):
            sheet.cell(row=num*4 + i, column=2).value = rec
            i += 1


        sheet.cell(row=num*4 + 1, column=3).value = 'col'
        i = 2
        for rec in np.mean(statsCol, axis=0):
            sheet.cell(row=num*4 + i, column=3).value = rec
            i += 1


        sheet.cell(row=num*4 + 1, column=4).value = 'sym'
        i = 2
        for rec in np.mean(statsSym, axis=0):
            sheet.cell(row=num*4 + i, column=4).value = rec
            i += 1


        sheet.cell(row=num*4 + 1, column=5).value = 'diag'
        i = 2
        for rec in np.mean(statsDiag, axis=0):
            sheet.cell(row=num*4 + i, column=5).value = rec
            i += 1


    wb.save(filename = destFile)