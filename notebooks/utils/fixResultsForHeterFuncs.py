import numpy as np
import rpy2.robjects as robjects
import rpy2.robjects.numpy2ri
from rpy2.robjects.packages import importr
import openpyxl
import pandas as pd

rpy2.robjects.numpy2ri.activate()
utils = importr('utils')
utils.chooseCRANmirror(ind=1)
utils.install_packages('Rssa')
rssa = importr('Rssa')


def findOvercomingMeanMax(ser, Q, tail, num, sheet, typeF='row'):
    maxVal = sheet[typeF][num*4 + 0]
    breakNum = None
    for i in range(Q-tail, len(ser)):
        if round(ser[i], 15) > round(maxVal, 15):
            breakNum = i + tail
            return [breakNum, ser[breakNum], ser[breakNum+10], ser[breakNum+20], ser[breakNum+30]]
    return [None, None, None, None, None]
        
def findOvercoming95Procentile(ser, Q, tail, num, sheet, typeF='row'):
    maxVal = sheet[typeF][num*4 + 1]
    breakNum = None
    for i in range(Q-tail, len(ser)):
        if round(ser[i], 15) > round(maxVal, 15):
            breakNum = i + tail
            return [breakNum, ser[breakNum], ser[breakNum+10], ser[breakNum+20], ser[breakNum+30]]
    return [None, None, None, None, None]
        
        
def rateOfIncrease(hm, Q, num, sheet, typeInc='meanMax'):
    if typeInc == 'meanMax':
        res = {
            'Row': findOvercomingMeanMax(hm.getRow(), Q, hm.T, num, sheet, 'row'),
            'Col': findOvercomingMeanMax(hm.getCol(), Q, hm.B, num, sheet, 'col'),
            'Sym': findOvercomingMeanMax(hm.getSym(), Q, hm.T, num, sheet, 'sym'),
            'Diag': findOvercomingMeanMax(hm.getDiag(), Q, hm.B + hm.T + 1, num, sheet, 'diag')
        }
    
    if typeInc == '95':
        res = {
            'Row': findOvercoming95Procentile(hm.getRow(), Q, hm.T, num, sheet, 'row'),
            'Col': findOvercoming95Procentile(hm.getCol(), Q, hm.B, num, sheet, 'col'),
            'Sym': findOvercoming95Procentile(hm.getSym(), Q, hm.T, num, sheet, 'sym'),
            'Diag': findOvercoming95Procentile(hm.getDiag(), Q, hm.B + hm.T + 1, num, sheet, 'diag')
        }

    return res


def insertRecord(sheet, num, colShift, valueMeanMax, value95):
    
    sheet.cell(row=num*9 + 4, column=1).value = 'Values of a series with an interval of 10'

    # Process the functions with meanMax
    for j, typeV in enumerate(valueMeanMax.keys()):
        sheet.cell(row=num*9 + 2, column=j*4+1).value = typeV
        sheet.cell(row=num*9 + 2, column=j*4+2).value = 'meanMax'
        sheet.cell(row=num*9 + 2, column=j*4+3).value = '95 procentile'
        sheet.cell(row=num*9 + 3, column=j*4+1).value = 'Point of overcoming'

        sheet.cell(row=num*9 + 4, column=j*4+1).value = 'X[point]'
        sheet.cell(row=num*9 + 5, column=j*4+1).value = 'X[point+10]'
        sheet.cell(row=num*9 + 6, column=j*4+1).value = 'X[point+20]'
        sheet.cell(row=num*9 + 7, column=j*4+1).value = 'X[point+30]'


        i = 3
        for rec in valueMeanMax[typeV]:
            sheet.cell(row=num*9 + i, column=j*4+2).value = rec
            i += 1

        # Process the functions with 95 procentile
        i = 3
        for rec in value95[typeV]:
            sheet.cell(row=num*9 + i, column=j*4+3).value = rec
            i += 1


def fixResults(dictHM:dict, Q, destFile:str, modellingResultsPath:str, title:str):
    '''
    :param dict dictHM: The dictionary where key is the type of heterogeneity matrix and value is a matrix. Example: { 'Permanent': hmPerm }.
    :param int Q: The point of perturbation.
    :param str destFile: Name of the file for saving results.
    :param str modellingResultsPath: Name of the file with modelling results.
    :param str title: Name of the worksheet.
    '''

    try:
        wb = openpyxl.load_workbook(filename = destFile)
        sheet = wb[title]
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = title
        sheet = wb[title]
    except KeyError:
        wb = openpyxl.load_workbook(filename = destFile)
        sheet = wb.create_sheet(title=title)

    modellingResults = pd.read_excel(modellingResultsPath, sheet_name='Modelling', engine='openpyxl')

    for num, typeV in enumerate(dictHM.keys()):
        
        hm = dictHM[typeV]

        sheet.cell(row=num*9 + 1, column=1).value = typeV

        rateOfIncreaseResMeanMax = rateOfIncrease(hm, Q, num, modellingResults, 'meanMax')
        rateOfIncreaseRes95 = rateOfIncrease(hm, Q, num, modellingResults, '95')

        insertRecord(sheet, num, 3, rateOfIncreaseResMeanMax, rateOfIncreaseRes95)


    wb.save(filename = destFile)