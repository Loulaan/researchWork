from typing import List

import numpy as np
import openpyxl
import pandas as pd

from utils.hmatr import Hmatr
from utils.modelling import insertRecord
from utils.utils import get_confidence_interval


def measure_statistics(func, Q, tail):
    # tail нужен для синхронизации функции разладки со значениями ряда. Q - точка возмущения в ряде, поэтому отнимаем для получения правильного значения функции разладки в точке возмущения.
    return np.max(func[:(Q - tail)]), np.quantile(func[:(Q - tail)], 0.95)


def get_modelling_statistics(hm, Q):
    """

    :param hm: Hmatr
    :param Q: Moment of perturbation
    :return: dict of lists with [max, 95proc] for every heterogeneity functions
    """
    return dict(
        zip(
            [
                'Row',
                'Col',
                'Sym',
                'Diag'
            ],
            [
                measure_statistics(hm.getRow(), Q, hm.T),
                measure_statistics(hm.getCol(), Q, hm.B),
                measure_statistics(hm.getSym(), Q, hm.T),
                measure_statistics(hm.getDiag(), Q, hm.B + hm.T + 1)
            ]
        )
    )


def find_perturbation_estimated_moment(series, value, Q, tail):
    breakNum = None
    for i in range(Q - tail, len(series)):
        if round(series[i], 10) > round(value, 10):
            breakNum = i + tail
            return [breakNum, series[i], series[Q - tail], series[Q - tail + 10], series[Q - tail + 20], series[Q - tail + 30]]
    return [None, None, None, None, None, None]


def findOvercomingStatistics(series, Q, tail, statistics: List):
    # tail нужен для синхронизации функции разладки со значениями ряда. Q - точка возмущения в ряде, поэтому отнимаем
    # для получения правильного значения функции разладки в точке возмущения.
    # Ищем точку преодоления среднего максимума (моделируемая характеристика при реализациях шума)
    maxVal = statistics[0]
    max_stats = find_perturbation_estimated_moment(series, maxVal, Q, tail)

    proc = statistics[1]
    proc_stats = find_perturbation_estimated_moment(series, proc, Q, tail)
    return [max_stats, proc_stats]


def rateOfIncrease(hm, Q: int, statistics: dict):
    """
    :param hm:
    :param Q:
    :param statistics: dict of lists with estimation of [max, 95 percentile] before moment Q
    :return: dict with lists of lists with [breakNum, ser[Q_], series[Q], series[Q + 10], series[Q + 20],
    series[Q + 30]] for every type of heterogeneity funcs
    """
    return dict(
        zip(
            [
                'Row',
                'Col',
                'Sym',
                'Diag'
            ],
            [
                findOvercomingStatistics(hm.getRow(), Q, hm.T, statistics['Row']),
                findOvercomingStatistics(hm.getCol(), Q, hm.B, statistics['Col']),
                findOvercomingStatistics(hm.getSym(), Q, hm.T, statistics['Sym']),
                findOvercomingStatistics(hm.getDiag(), Q, hm.B + hm.T + 1, statistics['Diag']),
            ]
        )
    )


def log_statistics(typeV, serieses, rows, cols, syms, diags, modelled_statistics_before_q):
    filename = f'tables/logs/{typeV}_'
    pd.DataFrame(serieses).to_csv(filename+'series.csv')
    pd.DataFrame(rows).to_csv(filename+'rows.csv')
    pd.DataFrame(cols).to_csv(filename+'cols.csv')
    pd.DataFrame(syms).to_csv(filename+'syms.csv')
    pd.DataFrame(diags).to_csv(filename+'diags.csv')
    pd.DataFrame(modelled_statistics_before_q).to_csv(filename+'modelled_statistics.csv')


def modelling_series_statistics(dictSeries: dict, iterNum: int, N: int, B: int, T: int, Q: int, L: int, r: int,
                              method: str, destFile: str, title: str, vareps: float):
    '''
    Modelling for series with noise
    :param dict dictSeries: The dictionary where key is the type of series and value is a series. Example: { 'Permanent': [x_1, ..., x_N] }.
    :param int iterNum: Number of iterations for modelling.
    :param int N: The len of original series.
    :param int B: The len of base subseries.
    :param int T: The len of test subseries.
    :param int Q: The point of perturbation.
    :param int L: The window len.
    :param int r: Number of eigen vectors.
    :param str method: SVD method.
    :param str destFile: Name of the file for saving results.
    :param str modellingResultsPath: Name of the file with modelling results.
    :param str title: Sheet title where results will be stored.
    :param float vareps: Variance of the noise.
    '''

    try:
        wb = openpyxl.load_workbook(filename=destFile)
        sheet = wb[title]
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = title
        sheet = wb[title]
    except KeyError:
        wb = openpyxl.load_workbook(filename=destFile)
        sheet = wb.create_sheet(title=title)

    confidence_intervals = dict()

    np.random.seed(12345)
    for num, typeV in enumerate(dictSeries.keys()):

        stats = []
        # FOR LOGGING
        serieses = []
        rows = []
        cols = []
        syms = []
        diags = []
        modelled_statistics_before_q = []

        for i in range(iterNum):
            eps = np.random.normal(scale=vareps, size=N) if typeV != 'Temporary' else \
                np.random.normal(scale=vareps / 2, size=N)

            seriesNoise = dictSeries[typeV] + eps

            serieses.append(seriesNoise)

            hm = Hmatr(seriesNoise, B, T, L, neig=r, svdMethod=method)

            rows.append(hm.getRow())
            cols.append(hm.getCol())
            syms.append(hm.getSym())
            diags.append(hm.getDiag())

            statistics = get_modelling_statistics(hm, Q)

            modelled_statistics_before_q.append(statistics)

            overcoming_stats = rateOfIncrease(hm, Q, statistics)
            stats.append(list(overcoming_stats.values()))

        log_statistics(typeV, serieses, rows, cols, syms, diags, modelled_statistics_before_q)

        statsMeanMax = np.asarray(stats)[:, :, 0, :]  # iters, typeV, type_stat, stats
        stats95 = np.asarray(stats)[:, :, 0, :]

        # Process results
        resMeanMax = dict()
        conf_ints_meanmax = dict()

        res95 = dict()
        conf_ints_95 = dict()
        # Get mean values of points of overcome of detection functions, num that points and values of [Q, Q+10, Q+20, Q+30]
        for idx, typeH in enumerate(['Row', 'Col', 'Sym', 'Diag']):
            # return statsMeanMax
            res, confs = get_statistics_for_detection_function_for_series_with_noise(stats=statsMeanMax, col_num=idx)
            resMeanMax[typeH] = res
            conf_ints_meanmax[typeH] = confs

            res, confs = get_statistics_for_detection_function_for_series_with_noise(stats=stats95, col_num=idx)
            res95[typeH] = res
            conf_ints_95[typeH] = confs

        sheet.cell(row=num * 10 + 1, column=1).value = typeV
        insertRecord(sheet, num, resMeanMax, res95)
        confidence_intervals[typeV] = dict()
        confidence_intervals[typeV]['mean_max'] = conf_ints_meanmax
        confidence_intervals[typeV]['95'] = conf_ints_95
    wb.save(filename=destFile)
    return confidence_intervals


def get_statistics_for_detection_function_for_series_with_noise(stats, col_num):
    # Вычленяем из общего набора статистик нужную нам функцию обнаружения и формируем средний результат для генерации таблицы.
    ans = []
    conf_intervals = dict()
    for i in range(6):
        if i == 0:
            tmp = stats[:, col_num, i]
            tmp = tmp[tmp != np.array(None)]
            ans.append(len(tmp))
            # Средняя точка преодоления промоделированного значения
            if len(tmp) == 0:
                ans.append(None)
            else:
                ans.append(round(np.mean(tmp), 3))
                conf_intervals['overc_p'] = get_confidence_interval(tmp, len(tmp))
        else:
            # Добавление средних значений для элементов ряда под номерами [Q_det, Q, Q+10, Q+20, Q+30]
            tmp = stats[:, col_num, i]
            if len(tmp[tmp != np.array(None)]) == 0:
                tmp = None
            else:
                tmp = tmp[tmp != np.array(None)]
                if tmp is not None:
                    if i == 1:
                        conf_intervals['overc_p_val'] = get_confidence_interval(tmp, len(tmp))
                tmp = np.mean(tmp)
            ans.append(tmp)
    return ans, conf_intervals
