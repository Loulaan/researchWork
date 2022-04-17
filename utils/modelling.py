from collections import namedtuple

import numpy as np
import rpy2.robjects as robjects
import rpy2.robjects.numpy2ri
from rpy2.robjects.packages import importr
import openpyxl
import pandas as pd
from scipy.stats import norm

from utils.hmatr import Hmatr
from utils.utils import get_confidence_interval

rpy2.robjects.numpy2ri.activate()
utils = importr('utils')
utils.chooseCRANmirror(ind=1)
# utils.install_packages('Rssa')
rssa = importr('Rssa')


def measureStatistics(func, Q, tail):
    # tail нужен для синхронизации функции разладки со значениями ряда. Q - точка возмущения в ряде, поэтому отнимаем для получения правильного значения функции разладки в точке возмущения. 
    return np.max(func[:(Q-tail)]), np.quantile(func[:(Q-tail)], 0.95)


def modellingNoiseStatistics(dictSeries:dict, iterNum:int, N:int, B:int, T:int, Q:int, L:int, r:int, method:str, vareps:float):
    """
        Моделирование статистик ряда (средний 95й процентиль и средний максимум) при различных реализациях шума до момента разладки методом Монте-Карло.
        Внимание, шум добавляется внутри метода!
        :param dict dictSeries: The dictionary where key is the type of series and value is a series. Example: { 'Permanent': [x_1, ..., x_N] }.
        :param int iterNum: Number of iterations for modelling.
        :param int N: The len of series.
        :param int B: The len of base subseries.
        :param int T: The len of test subseries.
        :param int Q: The point of perturbation.
        :param int L: The window len.
        :param int r: Number of eigen vectors.
        :param str method: SVD method.
        :param float vareps: Variance of the noise.
        :return: Pandas DataFrame
    """
    ds = pd.DataFrame(columns=["HeterType", "StatType", "row", "col", "sym", "diag"])
    ds["HeterType"] = ['Permanent', 'Permanent', 'Temporary', 'Temporary', 'Shifted', 'Shifted', 'Outlier', 'Outlier']
    ds["StatType"] = ["meanMax", "mean95Procentile"]*4
    
    row, col, sym, diag = [], [], [], []
    conf_intervals = namedtuple('c_intervals', 'Row Col Sym Diag')
    summary_conf_ints = dict()
    for typeV in dictSeries.keys():
        statsRow, statsCol, statsSym, statsDiag = [], [], [], []
        for i in range(iterNum):
            eps = np.random.randn(N) * vareps**2 if typeV != 'Temporary' else np.random.randn(N) * vareps**2/2
            
            seriesNoise = dictSeries[typeV] + eps
            hm = Hmatr(seriesNoise, B, T, L, neig=r, svdMethod=method)
            statsRow.append(measureStatistics(hm.getRow(), Q, hm.T))
            statsCol.append(measureStatistics(hm.getCol(), Q, hm.B))
            statsSym.append(measureStatistics(hm.getSym(), Q, hm.T))
            statsDiag.append(measureStatistics(hm.getDiag(), Q,  hm.B + hm.T + 1))
        row.append(np.mean(statsRow, axis=0))
        col.append(np.mean(statsCol, axis=0))
        sym.append(np.mean(statsSym, axis=0))
        diag.append(np.mean(statsDiag, axis=0))

        summary_conf_ints[typeV] = conf_intervals(
            get_confidence_interval(statsRow, iterNum),
            get_confidence_interval(statsCol, iterNum),
            get_confidence_interval(statsSym, iterNum),
            get_confidence_interval(statsDiag, iterNum)
        )

    print(summary_conf_ints)
    ds["row"] = np.array(row).reshape(-1, 1)
    ds["col"] = np.array(col).reshape(-1, 1)
    ds["sym"] = np.array(sym).reshape(-1, 1)
    ds["diag"] = np.array(diag).reshape(-1, 1)
    
    return ds, summary_conf_ints

def insert_cell(sheet, func_type, row_num, col_num, stats):
    # Вставляем значения статистик в таблицу
    sheet.cell(row=row_num, column=col_num).value = func_type
    i = 2
    for rec in np.mean(stats, axis=0):
        sheet.cell(row=row_num + i, column=col_num).value = rec
        i += 1

def findOvercomingMeanMax(ser, Q, tail, num, sheet, typeF='row'):
    # tail нужен для синхронизации функции разладки со значениями ряда. Q - точка возмущения в ряде, поэтому отнимаем
    # для получения правильного значения функции разладки в точке возмущения.
    # Ищем точку преодоления среднего максимума (моделируемая характеристика при реализациях шума)
    maxVal = sheet[typeF][num*2 + 0]
    breakNum = None
    for i in range(Q-tail, len(ser)):
        if round(ser[i], 10) > round(maxVal, 10):
            breakNum = i + tail
            return [breakNum, ser[i], ser[Q-tail], ser[Q-tail+10], ser[Q-tail+20], ser[Q-tail+30]]
    return [None, None, None, None, None, None]
        
def findOvercoming95Procentile(ser, Q, tail, num, sheet, typeF='row'):
    # tail нужен для синхронизации функции разладки со значениями ряда. Q - точка возмущения в ряде, поэтому отнимаем
    # для получения правильного значения функции разладки в точке возмущения.
    # Ищем точку преодоления среднего 95го процентиля (моделируемая характеристика при реализациях шума)
    maxVal = sheet[typeF][num*2 + 1]
    breakNum = None
    for i in range(Q-tail, len(ser)):
        if round(ser[i], 10) > round(maxVal, 10):
            breakNum = i + tail
            return [breakNum, ser[i], ser[Q-tail], ser[Q-tail+10], ser[Q-tail+20], ser[Q-tail+30]]
    return [None, None, None, None, None, None]
        
        
def rateOfIncrease(hm, Q, num, sheet, typeInc='meanMax'):
    if typeInc == 'meanMax':
        res = {
            'Row': findOvercomingMeanMax(hm.getRow(), Q, hm.T, num, sheet, 'row'),
            'Col': findOvercomingMeanMax(hm.getCol(), Q, hm.B, num, sheet, 'col'),
            'Sym': findOvercomingMeanMax(hm.getSym(), Q, hm.T, num, sheet, 'sym'),
            'Diag': findOvercomingMeanMax(hm.getDiag(), Q, hm.B + hm.T + 1, num, sheet, 'diag')
        }
    
    if typeInc == '95':
        res = {
            'Row': findOvercoming95Procentile(hm.getRow(), Q, hm.T, num, sheet, 'row'),
            'Col': findOvercoming95Procentile(hm.getCol(), Q, hm.B, num, sheet, 'col'),
            'Sym': findOvercoming95Procentile(hm.getSym(), Q, hm.T, num, sheet, 'sym'),
            'Diag': findOvercoming95Procentile(hm.getDiag(), Q, hm.B + hm.T + 1, num, sheet, 'diag')
        }
    return res



def insertRecord(sheet, num, valueMeanMax, value95, noise=True):

    multipRow = 10
    multipCol = 5

    for j, typeV in enumerate(valueMeanMax.keys()):
        # Insert info cells
        sheet.cell(row=num*multipRow + 2, column=j*multipCol+1).value = typeV
        sheet.cell(row=num*multipRow + 2, column=j*multipCol+2).value = 'meanMax'
        sheet.cell(row=num*multipRow + 2, column=j*multipCol+3).value = '95 procentile'

        if noise:
            sheet.cell(row=num*multipRow + 3, column=j*multipCol+1).value = 'Num Points of overcoming'
        sheet.cell(row=num*multipRow + 4 - int(not noise), column=j*multipCol+1).value = 'detected Q'

        sheet.cell(row=num*multipRow + 5 - int(not noise), column=j*multipCol+1).value = 'X[detected Q]'
        sheet.cell(row=num*multipRow + 6 - int(not noise), column=j*multipCol+1).value = 'X[Q]'
        sheet.cell(row=num*multipRow + 7 - int(not noise), column=j*multipCol+1).value = 'X[Q+10]'
        sheet.cell(row=num*multipRow + 8 - int(not noise), column=j*multipCol+1).value = 'X[Q+20]'
        sheet.cell(row=num*multipRow + 9 - int(not noise), column=j*multipCol+1).value = 'X[Q+30]'

        # Process the functions with meanMax
        i = 3
        for rec in valueMeanMax[typeV]:
            sheet.cell(row=num*multipRow + i, column=j*multipCol+2).value = rec
            i += 1

        # Process the functions with 95 procentile
        i = 3
        for rec in value95[typeV]:
            sheet.cell(row=num*multipRow + i, column=j*multipCol+3).value = rec
            i += 1


def modellingSeriesStatistics(dictSeries:dict, iterNum:int, N:int, B:int, T:int, Q:int, L:int, r:int, method:str, destFile:str, modellingResultsPath:str, title:str, vareps:float):
    '''
    Modelling for series with noise
    :param dict dictSeries: The dictionary where key is the type of series and value is a series. Example: { 'Permanent': [x_1, ..., x_N] }.
    :param int iterNum: Number of iterations for modelling.
    :param int N: The len of original series.
    :param int B: The len of base subseries.
    :param int T: The len of test subseries.
    :param int Q: The point of perturbation.
    :param int L: The window len.
    :param int r: Number of eigen vectors.
    :param str method: SVD method.
    :param str destFile: Name of the file for saving results.
    :param str modellingResultsPath: Name of the file with modelling results.
    :param str title: Sheet title where results will be stored.
    :param float vareps: Variance of the noise.
    '''

    try:
        wb = openpyxl.load_workbook(filename = destFile)
        sheet = wb[title]
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = title
        sheet = wb[title]
    except KeyError:
        wb = openpyxl.load_workbook(filename=destFile)
        sheet = wb.create_sheet(title=title)

    modellingResults = pd.read_csv(modellingResultsPath)
    confidence_intervals = dict()
    for num, typeV in enumerate(dictSeries.keys()):

        statsMeanMax = []
        stats95 = []

        for i in range(iterNum):
            eps = np.random.randn(N) * vareps**2 if typeV != 'Temporary' else np.random.randn(N) * vareps**2/2

            seriesNoise = dictSeries[typeV] + eps
            hm = Hmatr(seriesNoise, B, T, L, neig=r, svdMethod=method)
            statsMeanMax.append(list(rateOfIncrease(hm, Q, num, modellingResults, 'meanMax').values()))
            stats95.append(list(rateOfIncrease(hm, Q, num, modellingResults, '95').values()))

        statsMeanMax = np.array(statsMeanMax)
        stats95 = np.array(stats95)

        # Process results
        resMeanMax = dict()
        conf_ints_meanmax = dict()

        res95 = dict()
        conf_ints_95 = dict()
        # Get mean values of points of overcome of detection functions, num that points and values of [Q, Q+10, Q+20, Q+30]
        for idx, typeH in enumerate(['Row', 'Col', 'Sym', 'Diag']):
            res, confs = get_statistics_for_detection_function_for_series_with_noise(stats=statsMeanMax, col_num=idx)
            resMeanMax[typeH] = res
            conf_ints_meanmax[typeH] = confs

            res, confs = get_statistics_for_detection_function_for_series_with_noise(stats=stats95, col_num=idx)
            res95[typeH] = res
            conf_ints_95[typeH] = confs

        sheet.cell(row=num*10 + 1, column=1).value = typeV
        insertRecord(sheet, num, resMeanMax, res95)
        confidence_intervals[typeV] = dict()
        confidence_intervals[typeV]['mean_max'] = conf_ints_meanmax
        confidence_intervals[typeV]['95'] = conf_ints_95
    wb.save(filename=destFile)
    return confidence_intervals


def get_statistics_for_detection_function_for_series_with_noise(stats, col_num):
    # Вычленяем из общего набора статистик нужную нам функцию обнаружения и формируем средний результат для генерации таблицы.
    ans = []
    conf_intervals = dict()
    for i in range(6):
        if i == 0:
            tmp = stats[:, col_num, i]
            tmp = tmp[tmp != np.array(None)]
            ans.append(len(tmp))
            # Средняя точка преодоления промоделированного значения
            if len(tmp) == 0:
                ans.append(None)
            else:
                ans.append(round(np.mean(tmp), 3))
                conf_intervals['overc_p'] = get_confidence_interval(tmp, len(tmp))
        else:
            # Добавление средних значений для элементов ряда под номерами [Q_det, Q, Q+10, Q+20, Q+30]
            tmp = stats[:, col_num, i]
            if len(tmp[tmp != np.array(None)]) == 0:
                tmp = None
            else:
                tmp = tmp[tmp != np.array(None)]
            if i == 1:
                conf_intervals['overc_p_val'] = get_confidence_interval(tmp, len(tmp))
            tmp = np.mean(tmp)
            ans.append(tmp)
    return ans, conf_intervals



def fixSeriesStatistics(dictSeries:dict, B:int, T:int, Q:int, L:int, r:int, method:str, destFile:str, modellingResultsPath:str, title:str):
    '''
    Save results (statistics) for series without noise: point where value of detection function greater than modelled values.
    :param dict dictSeries: The dictionary where key is the type of series and value is a series. Example: { 'Permanent': [x_1, ..., x_N] }.
    :param int B: The len of base subseries.
    :param int T: The len of test subseries.
    :param int Q: The point of perturbation.
    :param int L: The window len.
    :param int r: Number of eigen vectors.
    :param str method: SVD method.
    :param str destFile: Name of the file for saving results.
    :param str modellingResultsPath: Name of the file with modelling results.
    :param str title: Sheet title where results will be stored.
    '''

    try:
        wb = openpyxl.load_workbook(filename = destFile)
        sheet = wb[title]
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = title
        sheet = wb[title]
    except KeyError:
        wb = openpyxl.load_workbook(filename = destFile)
        sheet = wb.create_sheet(title=title)

    modellingResults = pd.read_excel(modellingResultsPath, sheet_name='Modelling', engine='openpyxl')

    for num, typeV in enumerate(dictSeries.keys()):
        series = dictSeries[typeV]
        hm = Hmatr(series, B, T, L, neig=r, svdMethod=method)
        
        resMeanMax = rateOfIncrease(hm, Q, num, modellingResults, 'meanMax')
        res95 = rateOfIncrease(hm, Q, num, modellingResults, '95')

        sheet.cell(row=num*10 + 1, column=1).value = typeV
        insertRecord(sheet, num, resMeanMax, res95, False)

    
    wb.save(filename = destFile)


def findNonZeroPoint(ser, Q):
    maxVal = 0
    breakNum = None
    for i in range(len(ser)):
        if round(ser[i], 10) > round(maxVal, 10):
            breakNum = i
            return [breakNum, ser[Q], ser[Q+10], ser[Q+20], ser[Q+30]]
    return [None, None, None, None, None]


def rateOfIncreaseNonModelling(hm, Q):
    return {
        'Row': findNonZeroPoint(hm.getRow(sync=True), Q),
        'Col': findNonZeroPoint(hm.getCol(sync=True), Q),
        'Sym': findNonZeroPoint(hm.getSym(sync=True), Q),
        'Diag': findNonZeroPoint(hm.getDiag(sync=True), Q)
    }
    

def insert(sheet, num, values):

    multipRow = 9
    multipCol = 3

    for j, typeV in enumerate(values.keys()):
        # Insert info cells
        sheet.cell(row=num*multipRow + 2, column=j*multipCol+1).value = typeV
        sheet.cell(row=num*multipRow + 2, column=j*multipCol+2).value = 'ZeroValue'
        sheet.cell(row=num*multipRow + 3, column=j*multipCol+1).value = 'detected Q'
        sheet.cell(row=num*multipRow + 4, column=j*multipCol+1).value = 'X[Q]'
        sheet.cell(row=num*multipRow + 5, column=j*multipCol+1).value = 'X[Q+10]'
        sheet.cell(row=num*multipRow + 6, column=j*multipCol+1).value = 'X[Q+20]'
        sheet.cell(row=num*multipRow + 7, column=j*multipCol+1).value = 'X[Q+30]'

        # Process the functions
        i = 3
        # print(values[typeV])
        for rec in values[typeV]:
            sheet.cell(row=num*multipRow + i, column=j*multipCol+2).value = rec
            i += 1


def find_points_whith_non_zero_value(dictSeries:dict, B:int, T:int, Q:int, L:int, r:int, method:str, destFile:str, title:str):
    '''
    Save results (statistics) for series without noise: point where value of detection function greater than 0.
    :param dict dictSeries: The dictionary where key is the type of series and value is a series. Example: { 'Permanent': [x_1, ..., x_N] }.
    :param int B: The len of base subseries.
    :param int T: The len of test subseries.
    :param int Q: The point of perturbation.
    :param int L: The window len.
    :param int r: Number of eigen vectors.
    :param str method: SVD method.
    :param str destFile: Name of the file for saving results.
    :param str title: Sheet title where results will be stored.
    '''

    try:
        wb = openpyxl.load_workbook(filename = destFile)
        sheet = wb[title]
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = title
        sheet = wb[title]
    except KeyError:
        wb = openpyxl.load_workbook(filename = destFile)
        sheet = wb.create_sheet(title=title)

    for num, typeV in enumerate(dictSeries.keys()):
        series = dictSeries[typeV]
        hm = Hmatr(series, B, T, L, neig=r, svdMethod=method)
        
        res = rateOfIncreaseNonModelling(hm, Q)

        sheet.cell(row=num*9 + 1, column=1).value = typeV
        insert(sheet, num, res)

    
    wb.save(filename = destFile)
